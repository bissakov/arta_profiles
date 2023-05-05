from typing import Any, Dict 

import rich
from openpyxl import Workbook, styles
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def get_active_worksheet(wb: Workbook) -> Worksheet or None:
    return wb.worksheets[0]


def get_side(style: str = 'thin', color: str = '000000'):
    return Side(style=style, color=color)


def write_data(worksheet: Worksheet, family: Dict):
    for i, (header, data) in enumerate(family.items(), start=2):
        header_cell = worksheet.cell(column=i, row=1)
        header_cell.value = header
        header_cell.alignment = styles.Alignment(vertical='top', horizontal='center', wrap_text=True)
        header_cell.font = styles.Font(bold=True)
        header_cell.fill = styles.PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type = 'solid')
        header_cell.border = Border(bottom=get_side(style='thin'), left=get_side(style='thin'), right=get_side(style='thin'))

        data_cell = worksheet.cell(column=i, row=2)
        data_cell.value = data
        if isinstance(data, (int, float)):
            data_cell.alignment = styles.Alignment(vertical='center', horizontal='center', wrap_text=True)
        else:
            data_cell.alignment = styles.Alignment(vertical='center', wrap_text=True)
        data_cell.border = Border(bottom=get_side(style='thin'), left=get_side(style='thin'), right=get_side(style='thin'))


def set_column_width(worksheet: Worksheet, width: int):
    for i in range(2, worksheet.max_column + 1):
        if worksheet.cell(column=i, row=1).value is None:
            break
        worksheet.column_dimensions[get_column_letter(i)].width = width


def format_data(family: Dict) -> Dict:
    formatted_dict = dict()

    for key, val in family.items():
        if key == 'Члены семьи':
            for i, member in enumerate(val, start=1):
                formatted_dict[f'Член семьи {i}'] = f'{member["ИИН"]} {member["ФИО"]}'
        elif type(val) is dict:
            for key2, val2 in val.items():
                if key == 'Активы семьи' and val2 == 0:
                    continue    
                formatted_dict[key2] = val2 if type(val2) != float else round(val2, 2)
        elif type(val) is list:
            my_key = 'Рекомендация' if key == 'Рекомендации' else 'Риск'
            for i, el in enumerate(val, start=1):
                formatted_dict[f'{my_key} {i}'] = el

    return formatted_dict


def get_excel(family: Any = None) -> str:
    excel_name = f"excel/{family['Члены семьи'][0]['ИИН']}.xlsx"

   
    family = format_data({key: value for key, value in family.items() if family[key]})

    workbook = Workbook()
    worksheet = get_active_worksheet(wb=workbook)

    write_data(worksheet=worksheet, family=family)
    set_column_width(worksheet=worksheet, width=15)

    workbook.save(filename=excel_name)

    return excel_name

if __name__ == '__main__':
    get_excel()
